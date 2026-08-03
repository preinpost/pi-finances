#!/usr/bin/env python3
"""
parse-portal-excel.py — generate src/generated/apis.json from the official
KIS Open API portal Excel spec (전체 API 문서).

Source: https://apiportal.koreainvestment.com/files/download/apiCollection/API_COLLECTION
(다운로드: curl -L -o /tmp/kis_api_collection.xlsx ...)

Output schema (apis-v2):
{
  "generated": ISO timestamp,
  "source": "apiportal.koreainvestment.com 전체 API 규격 (Excel)",
  "excel": "API_COLLECTION",
  "apis": {
    "<category>.<key>": {
      "name", "category", "api_id", "kind" (REST|WEBSOCKET),
      "method", "api_path", "tr_id_real": [..], "tr_id_paper": [..],
      "description", "headers": {name: {name_kr,type,required,length,desc}},
      "query": {...}, "body": {...}, "response": {...},
      "example_request", "example_response"
    }
  }
}
"""
import json
import re
import sys
from datetime import datetime, timezone

import openpyxl

XLSX = sys.argv[1] if len(sys.argv) > 1 else "/tmp/kis_api_collection.xlsx"
OUT = sys.argv[2] if len(sys.argv) > 2 else "src/generated/apis.json"

# 메뉴 위치(카테고리) → API 키 카테고리 슬러그
CATEGORY_MAP = [
    ("[국내주식]", "domestic_stock"),
    ("[국내선물옵션]", "domestic_futureoption"),
    ("[해외주식]", "overseas_stock"),
    ("[해외선물옵션]", "overseas_futureoption"),
    ("[장내채권]", "domestic_bond"),
    ("OAuth", "oauth"),
]

def category_slug(menu: str) -> str:
    for prefix, slug in CATEGORY_MAP:
        if menu.startswith(prefix):
            return slug
    return "etc"

def norm(name: str) -> str:
    """시트명/API명 매칭용 정규화 (슬래시→언더스코어, 공백/특수문자 제거)"""
    return re.sub(r"[^0-9A-Za-z가-힣]", "", name.replace("/", "_"))

def parse_tr_ids(raw: str) -> list[str]:
    if not raw or "미지원" in raw or "없음" in raw:
        return []
    # "(매도) TTTC0011U (매수) TTTC0012U" → ["TTTC0011U", "TTTC0012U"]
    # TR_ID는 9~13자리 (예: CTRGA011R, HHDFS00000300, FHKST03010100)
    ids = re.findall(r"\b[A-Z][A-Z0-9]{8,12}\b", raw or "")
    return list(dict.fromkeys(ids))

def parse_fields(rows: list[list]) -> dict:
    """Layout 표(rows 16+)에서 섹션별 필드 파싱.
    열 A는 병합된 섹션 헤더(Request Header 등), 필드는 열 B(Element)부터 시작."""
    sections = {}
    cur = None
    for row in rows:
        cells = [str(c).strip() if c is not None else "" for c in row]
        hdr = cells[0]
        element = cells[1] if len(cells) > 1 else ""
        if hdr in ("Request Header", "Request Query Parameter", "Request Body",
                   "Response Header", "Response Body"):
            cur = hdr
            sections[cur] = {}
        if not cur or not element or element == "Element":
            continue
        sections[cur][element] = {
            "name_kr": cells[2] if len(cells) > 2 else "",
            "type": cells[3] if len(cells) > 3 else "",
            "required": cells[4] == "Y" if len(cells) > 4 else False,
            "length": cells[5] if len(cells) > 5 else "",
            "desc": cells[6] if len(cells) > 6 else "",
        }
    return sections

wb = openpyxl.load_workbook(XLSX, read_only=True)

# 1) API 목록 시트
list_ws = wb["API 목록"]
apis = {}
list_rows = list(list_ws.iter_rows(values_only=True))
header = [str(c) if c else "" for c in list_rows[0]]
col = {h: i for i, h in enumerate(header)}
n = 0
for row in list_rows[1:]:
    if not row or not any(row):
        continue
    get = lambda h: str(row[col[h]]).strip() if col.get(h) is not None and row[col[h]] is not None else ""
    api_name = get("API 명")
    menu = get("메뉴 위치")
    kind = get("API 통신방식")
    if not api_name:
        continue
    cat = category_slug(menu)
    # 시트명 매칭
    sheet = None
    target = norm(api_name)
    for sn in wb.sheetnames:
        if sn == "API 목록":
            continue
        if norm(sn) == target:
            sheet = sn
            break
    tr_real = parse_tr_ids(get("실전 TR_ID"))
    tr_paper = parse_tr_ids(get("모의 TR_ID"))
    method = get("HTTP Method")
    url = get("URL 명")
    api_id = get("API ID")
    # kind 교정: URL 기반이 목록 시트의 통신방식보다 신뢰성 높음
    # (목록 시트는 일부 항목을 잘못 분류 — 예: 국내주식-163 REST, 장운영정보 웹소켓)
    if url.startswith("/uapi/"):
        kind = "REST"
    elif url.startswith("/tryitout/") or url.startswith("ws") or "websocket" in url.lower():
        kind = "WEBSOCKET"
    key = f"{cat}.{api_id}" if api_id else f"{cat}.{re.sub(r'[^0-9A-Za-z가-힣]', '_', api_name)}"
    key = re.sub(r"[^\w가-힣.-]", "_", key)  # 공백/특수문자 → _ (키 안전화)
    entry = {
        "name": api_name,
        "category": menu,
        "api_id": api_id,
        "kind": kind,
        "method": method,
        "api_path": url,
        "tr_id_real": tr_real,
        "tr_id_paper": tr_paper,
        "sheet": sheet,
    }
    # 2) 상세 시트 파싱
    if sheet:
        ws = wb[sheet]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        # 개요
        for r in rows:
            if r and str(r[0]).strip() == "개요" and len(r) > 1 and r[1] is not None and str(r[1]).strip():
                entry["description"] = str(r[1]).strip()
                break
        # Layout 표 (구분 헤더 이후)
        try:
            layout_idx = next(i for i, r in enumerate(rows) if r and str(r[0]).strip() == "Layout")
            example_idx = next((i for i, r in enumerate(rows) if r and str(r[0]).strip() == "Example"), len(rows))
            sections = parse_fields(rows[layout_idx + 1:example_idx])
            entry["headers"] = sections.get("Request Header", {})
            entry["query"] = sections.get("Request Query Parameter", {})
            entry["body"] = sections.get("Request Body", {})
            entry["response_header"] = sections.get("Response Header", {})
            entry["response"] = sections.get("Response Body", {})
        except StopIteration:
            pass
        # 예제
        for i, r in enumerate(rows):
            if r and str(r[0]).strip() == "Request Example (Python)" and i + 1 < len(rows):
                entry["example_request"] = str(rows[i + 1][0]).strip() if rows[i + 1][0] else ""
            if r and str(r[0]).strip() == "Response Example" and i + 1 < len(rows):
                entry["example_response"] = str(rows[i + 1][0]).strip() if rows[i + 1][0] else ""
    apis[key] = entry
    n += 1

out = {
    "generated": datetime.now(timezone.utc).isoformat(),
    "source": "apiportal.koreainvestment.com 전체 API 규격 (Excel: API_COLLECTION)",
    "excel": "API_COLLECTION",
    "apis": apis,
}
with open(OUT, "w", encoding="utf-8") as f:
    json.dump(out, f, ensure_ascii=False, indent=1)
    f.write("\n")

kinds = {}
cats = {}
for e in apis.values():
    kinds[e["kind"]] = kinds.get(e["kind"], 0) + 1
    cats[e["category"]] = cats.get(e["category"], 0) + 1
with_sheet = sum(1 for e in apis.values() if e.get("sheet"))
with_tr = sum(1 for e in apis.values() if e.get("tr_id_real"))
print(f"wrote {OUT}")
print(f"  total APIs : {n}")
print(f"  kinds      : {kinds}")
print(f"  상세 시트  : {with_sheet}/{n} (파라미터/응답 명세 포함)")
print(f"  TR_ID 있음 : {with_tr}/{n}")
print(f"  카테고리   : {cats}")
